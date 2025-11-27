from flask import Flask, render_template, jsonify, request, send_file, Response, session, redirect, url_for
from werkzeug.security import check_password_hash, generate_password_hash
from functools import wraps
import json
from pathlib import Path
import io
from datetime import datetime
import sqlite3
import threading

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment


app = Flask(__name__, static_folder='static', template_folder='templates')
app.secret_key = 'your-secret-key-here-change-in-production'
DB_FILE = Path(__file__).parent / 'todo_app.db'

# Thread-local storage for database connections
local = threading.local()


def get_db():
	"""Get thread-local database connection"""
	if not hasattr(local, 'db'):
		local.db = sqlite3.connect(str(DB_FILE))
		local.db.row_factory = sqlite3.Row  # Return rows as dictionaries
	return local.db


def init_db():
	"""Initialize database schema"""
	db = get_db()
	cursor = db.cursor()

	# Create users table
	cursor.execute('''
		CREATE TABLE IF NOT EXISTS users (
			id INTEGER PRIMARY KEY AUTOINCREMENT,
			username TEXT UNIQUE NOT NULL,
			password TEXT NOT NULL,
			created_at TEXT NOT NULL
		)
	''')

	# Create tasks table
	cursor.execute('''
		CREATE TABLE IF NOT EXISTS tasks (
			id INTEGER PRIMARY KEY AUTOINCREMENT,
			text TEXT NOT NULL,
			important INTEGER NOT NULL DEFAULT 0,
			urgent INTEGER NOT NULL DEFAULT 0,
			completed INTEGER NOT NULL DEFAULT 0,
			user_id INTEGER NOT NULL,
			created_at TEXT,
			completed_at TEXT,
			deadline TEXT,
			FOREIGN KEY (user_id) REFERENCES users (id)
		)
	''')

	db.commit()


def add_dummy_data():
	"""Add dummy data for testing"""
	db = get_db()
	cursor = db.cursor()

	# Check if data already exists
	cursor.execute('SELECT COUNT(*) FROM users')
	if cursor.fetchone()[0] > 0:
		return  # Data already exists

	# Add dummy users
	# Password for all users is "password123"
	hashed_password = generate_password_hash('password123', method='scrypt')

	users = [
		('vicky', hashed_password, datetime.now().strftime('%Y-%m-%d %H:%M:%S')),
		('admin', hashed_password, datetime.now().strftime('%Y-%m-%d %H:%M:%S')),
		('testuser', hashed_password, datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
	]

	cursor.executemany('INSERT INTO users (username, password, created_at) VALUES (?, ?, ?)', users)

	# Add dummy tasks
	tasks = [
		('Complete project documentation', 1, 1, 0, 1, datetime.now().strftime('%Y-%m-%d %H:%M:%S'), None, '2025-11-30'),
		('Review code changes', 1, 0, 0, 1, datetime.now().strftime('%Y-%m-%d %H:%M:%S'), None, '2025-12-01'),
		('Team meeting preparation', 0, 1, 0, 1, datetime.now().strftime('%Y-%m-%d %H:%M:%S'), None, '2025-11-28'),
		('Update database schema', 0, 0, 1, 1, datetime.now().strftime('%Y-%m-%d %H:%M:%S'), datetime.now().strftime('%Y-%m-%d %H:%M:%S'), '2025-11-25'),
		('Setup production environment', 1, 1, 0, 2, datetime.now().strftime('%Y-%m-%d %H:%M:%S'), None, '2025-12-05'),
		('Write unit tests', 1, 0, 0, 2, datetime.now().strftime('%Y-%m-%d %H:%M:%S'), None, '2025-12-03'),
		('Fix login bug', 0, 1, 1, 2, datetime.now().strftime('%Y-%m-%d %H:%M:%S'), datetime.now().strftime('%Y-%m-%d %H:%M:%S'), '2025-11-26'),
		('Learn SQLite basics', 0, 0, 0, 3, datetime.now().strftime('%Y-%m-%d %H:%M:%S'), None, '2025-12-10')
	]

	cursor.executemany('''
		INSERT INTO tasks (text, important, urgent, completed, user_id, created_at, completed_at, deadline)
		VALUES (?, ?, ?, ?, ?, ?, ?, ?)
	''', tasks)

	db.commit()


@app.before_request
def before_request():
	"""Initialize database before first request"""
	if not hasattr(app, 'db_initialized'):
		init_db()
		add_dummy_data()
		app.db_initialized = True


@app.teardown_appcontext
def close_db(error):
	"""Close database connection at end of request"""
	if hasattr(local, 'db'):
		local.db.close()
		delattr(local, 'db')


def login_required(f):
	@wraps(f)
	def decorated_function(*args, **kwargs):
		if 'user_id' not in session:
			# For HTML routes, redirect to login
			if request.path in ['/', '/logout']:
				return redirect(url_for('login_page'))
			# For API routes, return 401
			return jsonify({'error': 'Unauthorized'}), 401
		return f(*args, **kwargs)
	return decorated_function


def row_to_dict(row):
	"""Convert SQLite row to dictionary"""
	return dict(row) if row else None


@app.route('/login', methods=['GET'])
def login_page():
	if 'user_id' in session:
		return redirect(url_for('index'))
	return render_template('login.html')


@app.route('/login', methods=['POST'])
def login():
	data = request.get_json(force=True)
	username = (data.get('username') or '').strip()
	password = data.get('password') or ''

	if not username or not password:
		return jsonify({'success': False, 'error': 'Username and password required'}), 400

	db = get_db()
	cursor = db.cursor()
	cursor.execute('SELECT * FROM users WHERE username = ?', (username,))
	user = cursor.fetchone()

	if user and check_password_hash(user['password'], password):
		session['user_id'] = user['id']
		session['username'] = user['username']
		return jsonify({'success': True})

	return jsonify({'success': False, 'error': 'Invalid username or password'}), 401


@app.route('/signup', methods=['GET'])
def signup_page():
	if 'user_id' in session:
		return redirect(url_for('index'))
	return render_template('signup.html')


@app.route('/signup', methods=['POST'])
def signup():
	data = request.get_json(force=True)
	username = (data.get('username') or '').strip()
	password = data.get('password') or ''

	if not username or not password:
		return jsonify({'success': False, 'error': 'Username and password required'}), 400

	if len(password) < 6:
		return jsonify({'success': False, 'error': 'Password must be at least 6 characters'}), 400

	db = get_db()
	cursor = db.cursor()

	# Check if username already exists
	cursor.execute('SELECT id FROM users WHERE username = ?', (username,))
	if cursor.fetchone():
		return jsonify({'success': False, 'error': 'Username already exists'}), 400

	# Create new user
	hashed_password = generate_password_hash(password, method='scrypt')
	created_at = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

	cursor.execute('''
		INSERT INTO users (username, password, created_at)
		VALUES (?, ?, ?)
	''', (username, hashed_password, created_at))
	db.commit()

	new_user_id = cursor.lastrowid

	# Auto-login
	session['user_id'] = new_user_id
	session['username'] = username

	return jsonify({'success': True})


@app.route('/logout')
@login_required
def logout():
	session.clear()
	return redirect(url_for('login_page'))


@app.route('/')
@login_required
def index():
	return render_template('index.html', username=session.get('username'))


@app.route('/api/tasks', methods=['GET'])
@login_required
def get_tasks():
	db = get_db()
	cursor = db.cursor()
	cursor.execute('''
		SELECT * FROM tasks
		WHERE user_id = ?
		ORDER BY id DESC
	''', (session['user_id'],))
	tasks = [row_to_dict(row) for row in cursor.fetchall()]
	return jsonify(tasks)


@app.route('/api/tasks', methods=['POST'])
@login_required
def add_task():
	data = request.get_json(force=True)
	text = (data.get('text') or '').strip()
	if not text:
		return jsonify({'error': 'empty task'}), 400

	important = 1 if data.get('important') else 0
	urgent = 1 if data.get('urgent') else 0
	deadline = data.get('deadline')
	created_at = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

	db = get_db()
	cursor = db.cursor()
	cursor.execute('''
		INSERT INTO tasks (text, important, urgent, completed, user_id, created_at, deadline)
		VALUES (?, ?, ?, 0, ?, ?, ?)
	''', (text, important, urgent, session['user_id'], created_at, deadline))
	db.commit()

	task_id = cursor.lastrowid
	cursor.execute('SELECT * FROM tasks WHERE id = ?', (task_id,))
	task = row_to_dict(cursor.fetchone())

	return jsonify(task), 201


@app.route('/api/tasks/<int:task_id>/toggle', methods=['POST'])
@login_required
def toggle_task(task_id):
	db = get_db()
	cursor = db.cursor()

	# Check if task exists and user owns it
	cursor.execute('SELECT * FROM tasks WHERE id = ? AND user_id = ?', (task_id, session['user_id']))
	task = cursor.fetchone()

	if not task:
		return jsonify({'error': 'not found'}), 404

	# Toggle completed status
	new_completed = 0 if task['completed'] else 1
	completed_at = datetime.now().strftime('%Y-%m-%d %H:%M:%S') if new_completed else None

	cursor.execute('''
		UPDATE tasks
		SET completed = ?, completed_at = ?
		WHERE id = ?
	''', (new_completed, completed_at, task_id))
	db.commit()

	# Return updated task
	cursor.execute('SELECT * FROM tasks WHERE id = ?', (task_id,))
	updated_task = row_to_dict(cursor.fetchone())
	return jsonify(updated_task)


@app.route('/api/tasks/<int:task_id>', methods=['DELETE'])
@login_required
def delete_task(task_id):
	db = get_db()
	cursor = db.cursor()

	# Check if task exists and user owns it
	cursor.execute('SELECT id FROM tasks WHERE id = ? AND user_id = ?', (task_id, session['user_id']))
	task = cursor.fetchone()

	if not task:
		return jsonify({'error': 'not found'}), 404

	# Delete task
	cursor.execute('DELETE FROM tasks WHERE id = ?', (task_id,))
	db.commit()

	return jsonify({'status': 'deleted'})


@app.route('/export/<which>')
@login_required
def export_xlsx(which):
	"""Export tasks as .xlsx. which: 'pending' or 'done' or 'all'"""
	if openpyxl is None:
		return jsonify({'error': 'openpyxl not installed on server'}), 500

	db = get_db()
	cursor = db.cursor()

	# Build query based on which type
	if which == 'pending':
		cursor.execute('SELECT * FROM tasks WHERE user_id = ? AND completed = 0', (session['user_id'],))
	elif which == 'done':
		cursor.execute('SELECT * FROM tasks WHERE user_id = ? AND completed = 1', (session['user_id'],))
	else:
		cursor.execute('SELECT * FROM tasks WHERE user_id = ?', (session['user_id'],))

	rows = [row_to_dict(row) for row in cursor.fetchall()]

	wb = openpyxl.Workbook()
	ws = wb.active
	ws.title = 'Tasks'

	header = ['ID', 'Task', 'Priority', 'Status']
	ws.append(header)

	# header style
	header_font = Font(bold=True, color='FFFFFF')
	header_fill = PatternFill(start_color='4F46E5', end_color='4F46E5', fill_type='solid')
	for col in range(1, len(header) + 1):
		cell = ws.cell(row=1, column=col)
		cell.font = header_font
		cell.fill = header_fill
		cell.alignment = Alignment(horizontal='center')

	# Priority fills
	fills = {
		'prio-critical': PatternFill(start_color='FFCCCC', end_color='FFCCCC', fill_type='solid'),
		'prio-urgent': PatternFill(start_color='FFF0B8', end_color='FFF0B8', fill_type='solid'),
		'prio-important': PatternFill(start_color='C8F0D0', end_color='C8F0D0', fill_type='solid'),
		'prio-normal': PatternFill(start_color='E9EAEC', end_color='E9EAEC', fill_type='solid'),
	}

	def priority_label(t):
		if t.get('important') and t.get('urgent'):
			return 'Important & Urgent'
		if not t.get('important') and t.get('urgent'):
			return 'NotImportant & Urgent'
		if t.get('important') and not t.get('urgent'):
			return 'Important & NotUrgent'
		return 'NotImportant & NotUrgent'

	for t in rows:
		pri = priority_label(t)
		status = 'Done' if t.get('completed') else 'Pending'
		# force string for very large numeric-looking text to avoid Excel scientific notation
		id_val = str(t.get('id', ''))
		task_text = str(t.get('text', ''))
		ws.append([id_val, task_text, pri, status])

	# Apply fills per-row based on priority
	for r in range(2, ws.max_row + 1):
		pri_val = ws.cell(row=r, column=3).value
		cls = 'prio-normal'
		if pri_val == 'Important & Urgent':
			cls = 'prio-critical'
		elif pri_val == 'NotImportant & Urgent':
			cls = 'prio-urgent'
		elif pri_val == 'Important & NotUrgent':
			cls = 'prio-important'
		fill = fills.get(cls)
		if fill:
			for c in range(1, 5):
				ws.cell(row=r, column=c).fill = fill
		# if done, make text slightly dim / italic
		if ws.cell(row=r, column=4).value == 'Done':
			for c in range(1, 5):
				cell = ws.cell(row=r, column=c)
				cell.font = Font(italic=True, color='444444')

	# column widths
	ws.column_dimensions['A'].width = 8
	ws.column_dimensions['B'].width = 50
	ws.column_dimensions['C'].width = 22
	ws.column_dimensions['D'].width = 12

	bio = io.BytesIO()
	wb.save(bio)
	bio.seek(0)

	filename = f"tasks_{which}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
	return send_file(bio, as_attachment=True, download_name=filename, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


if __name__ == '__main__':
	app.run(debug=True, port=5000)
